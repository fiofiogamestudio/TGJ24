import os
import json
from openpyxl import load_workbook

# 备份
import shutil

source_file = "dialog.xlsx"
backup_file = "dialog_backup.xlsx"

shutil.copy(source_file, backup_file)

SEQUENCE = "DialogSequece"
SELECTION = "DialogSelection"
PACKAGE = "DialogPackage"

SEQUENCE_KEYS = ["nextDialogID"]
SELECTION_KEYS = ["selectionList"]
PACKAGE_KEYS = ["targetInventoryID", "successNextDialogID", "failedNextDialogID", "successCallback", "failedCallback"]

def remove_keys(item, keys):
    for key in keys:
        if item[key] != "" and item[key] != None:
            item_id = item["dialogID"]
            item_type = item["$type"]
            print(f"第{item_id}个对话是{item_type}类型，存在多余键{key}，值为{item[key]}!")
        item.pop(key, None)
    return item

def reformat_sequence(item):
    return remove_keys(item, SELECTION_KEYS + PACKAGE_KEYS)

def reformat_selection(item):
    if "selectionList" in item and item["selectionList"]:
        item["selectionList"] = list(map(int, item["selectionList"].split(',')))
    return remove_keys(item, SEQUENCE_KEYS + PACKAGE_KEYS)

def reformat_package(item):
    return remove_keys(item,SEQUENCE_KEYS + SELECTION_KEYS)

# 打开Excel文件
wb = load_workbook("dialog.xlsx")
ws = wb.active

# 读取标题行
headers = [cell.value for cell in ws[1]]

# 从Excel表中读取数据并构建字典
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    item = {headers[col_num]: cell.value for col_num, cell in enumerate(row)}
    
    # 根据$type字段重新格式化数据
    item_type = item["$type"]
    if SEQUENCE in item_type:
        item = reformat_sequence(item)
    if SELECTION in item_type:
        item = reformat_selection(item)
    if PACKAGE in item_type:
        item = reformat_package(item)
    
    data.append(item)

# 将数据保存到JSON文件中
if os.path.exists("dialog.json"):
    os.remove("dialog.json")
with open("dialog.json", "w", encoding="utf-8") as file:
    json.dump({"dialogData": data}, file, ensure_ascii=False, indent=4)
