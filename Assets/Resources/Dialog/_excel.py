import os
import json
from openpyxl import Workbook

# 备份
import shutil

source_file = "dialog.json"
backup_file = "dialog_backup.json"

shutil.copy(source_file, backup_file)


# 不同类型数据备注
# 通用   dialogID    dialogType  holderID    content    dialogCallback    |   faceID
# 线性  nextDialogID    
# 选择  selectionList   
# 背包  targetInventoryID successNextDialogID failedNextDialogID  successCallback failedCallback

SEQUENCE = "DialogSequece"
SELECTION = "DialogSelection"
PACKAGE = "DialogPackage"

def process_sequence(item):
    return item

def process_selection(item):
    if "selectionList" in item:
        item["selectionList"] = ','.join(map(str, item["selectionList"]))
    return item

def process_package(item):
    return item

# 从"dialog.json"文件中加载JSON数据
with open("dialog.json", "r", encoding="utf-8") as file:
    raw_data = json.load(file)["dialogData"]

processed_data = []
for item in raw_data:
    item_type = item["$type"]
    if SEQUENCE in item_type:
        processed_data.append(process_sequence(item))
    if SELECTION in item_type:
        processed_data.append(process_selection(item))
    if PACKAGE in item_type:
        processed_data.append(process_package(item))


# 获取所有可能title 
# headers = set(key for item in processed_data for key in item.keys())
# 手动定义headers顺序
headers = [
    "$type", "dialogID", "dialogType", "holderID", "content", "dialogCallback", "faceID",
    "nextDialogID", "selectionList", 
    "targetInventoryID", "successNextDialogID", "failedNextDialogID", "successCallback", "failedCallback"
]


# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active

# 添加标题行
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# 填充数据
for row_num, item in enumerate(processed_data, 2):
    try:
        for col_num, key in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=item.get(key, ""))
    except Exception as e:
        content_preview = item.get("content", "")[:5]
        print(f"第{row_num}出错！预览：{content_preview}... Error{e}")

# 如果dialog.xlsx文件存在，删除它
if os.path.exists("dialog.xlsx"):
    os.remove("dialog.xlsx")
# 更新dialog.xlsx
wb.save("dialog.xlsx")
